from openpyxl import Workbook, load_workbook

# 新建工作簿
wb = Workbook()
# 选择默认的工作簿
sheet = wb.active
# 为当前工作表命名
sheet.title = '测试1表'

# 往A1单元格写入信息
sheet['b1'] = 'testAAA'
sheet['a2'] = 'testAAA'
sheet['b3'] = 'testAAA'

# 写入单行多个信息
row = ['111', '222', '3334']
sheet.append(row)

# 写入多行信息
data = [
    ['aaa', '333', 'asd'],
    ['zzz', '456', 'qaa'],
    ['sss', '65445', 'dh'],
    ['dsg', '888', 'jddf']
]
for i in data:
    sheet.append(i)

# 保存Excel
wb.save('测试excel.xlsx')

# 打开excel表
lw = load_workbook('测试excel.xlsx')
# 打印出所有工作表名称
print(lw.sheetnames)
# 选择工作表
show_sheet = lw['测试1表']
# 打印当前工作表A2单元格内容
print(show_sheet['A2'].value)

# 打印所有单元格的值
for row in show_sheet.rows:
    for i in row:
        print(i.value, end=' ')
    print()
